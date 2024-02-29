from tkinter import Tk
from tkinter.filedialog import askopenfilename
import openpyxl
import zipfile
import os

# Create a Tkinter root window (it won't be shown)
root = Tk()
root.withdraw()

# Ask the user to select an Excel file using Windows Explorer
excel_file_path = askopenfilename(filetypes=[("Excel Files", "*.xlsx;*.xlsm")])

# Check if the user selected a file
if excel_file_path:
    # Load the Excel workbook using openpyxl
    workbook = openpyxl.load_workbook(excel_file_path)
    
     # Get the number of worksheets in the workbook
    num_sheets = len(workbook.sheetnames)
    print(f"Number of sheets in the workbook: {num_sheets}")


    # Get the sheet names in the workbook
    sheet_names = workbook.sheetnames
    print(f"Sheet names in the workbook: {sheet_names}")
    
    # Check if each sheet is protected or not
    for sheet_name in sheet_names:
        sheet = workbook[sheet_name]
        if sheet.protection.sheet:
            print(f"{sheet_name} is protected.")
        else:
            print(f"{sheet_name} is not protected.")
            
# Check if the user selected a file
if excel_file_path:
    # Replace the extension with '.zip'
    zip_file_path = os.path.splitext(excel_file_path)[0] + '.zip'

    # Rename the file by moving it to a new path
    os.rename(excel_file_path, zip_file_path)

    print(f"Excel file '{excel_file_path}' has been renamed to '{zip_file_path}'.")
else:
    print("No Excel file selected.")
    
    


# Check if the user selected a file
if zip_file_path:
    # Create a folder with the same name as the zip file (excluding the extension)
    folder_name = os.path.splitext(os.path.basename(zip_file_path))[0]
    os.makedirs(folder_name, exist_ok=True)

    # Create a ZipFile object
    with zipfile.ZipFile(zip_file_path, 'r') as zip_file:
        # Extract the contents to the folder
        zip_file.extractall(folder_name)

    print(f"Contents of '{zip_file_path}' have been extracted to the folder '{folder_name}'.")
else:
    print("No zip file selected.")

            
