from tkinter import Tk
from tkinter.filedialog import askopenfilename
import openpyxl
import zipfile
import os

# Create a Tkinter root window (it won't be shown)
root = Tk()
root.withdraw()

# Ask the user to select an Excel file using Windows Explorer
excel_file_path = askopenfilename(filetypes=[("Excel Files", "*.xlsx;*.xlsm")])

# Check if the user selected a file
if excel_file_path:
    # Load the Excel workbook using openpyxl
    workbook = openpyxl.load_workbook(excel_file_path)
    
     # Get the number of worksheets in the workbook
    num_sheets = len(workbook.sheetnames)
    print(f"Number of sheets in the workbook: {num_sheets}")


    # Get the sheet names in the workbook
    sheet_names = workbook.sheetnames
    print(f"Sheet names in the workbook: {sheet_names}")
    
    # Check if each sheet is protected or not
    for sheet_name in sheet_names:
        sheet = workbook[sheet_name]
        if sheet.protection.sheet:
            print(f"{sheet_name} is protected.")
        else:
            print(f"{sheet_name} is not protected.")
            
            
# Check if the user selected a file
if excel_file_path:
    # Replace the extension with '.zip'
    zip_file_path = os.path.splitext(excel_file_path)[0] + '.zip'

    # Create a new zip file
    with zipfile.ZipFile(zip_file_path, 'w') as zip_file:
        # Add the Excel file to the zip file with the same name
        zip_file.write(excel_file_path, arcname=os.path.basename(excel_file_path))

    print(f"Excel file '{excel_file_path}' has been converted to '{zip_file_path}' successfully.")
    # Now you can work with the workbook as needed
    # For example, you can access sheets, read data, etc.

    # Close the workbook when done
    workbook.close()
else:
    print("No file selected.")
